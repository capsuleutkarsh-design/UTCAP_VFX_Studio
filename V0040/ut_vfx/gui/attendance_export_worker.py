"""Attendance Excel export worker extracted from AttendanceTab."""

from __future__ import annotations

import calendar
from datetime import datetime

from PySide6.QtCore import QThread, Signal


class ExcelExportWorker(QThread):
    """Background worker for openpyxl Excel export to prevent UI freezes."""

    finished_export = Signal(bool, str)  # success, message

    def __init__(self, path, year, month, users, data, late_hour, late_min):
        super().__init__()
        self.path = path
        self.year = year
        self.month = month
        self.users = users
        self.data = data
        self.late_hour = late_hour
        self.late_min = late_min

    def run(self):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            days_in_month = calendar.monthrange(self.year, self.month)[1]

            wb = Workbook()
            ws = wb.active
            ws.title = f"{calendar.month_name[self.month]} {self.year}"

            headers = ["Name", "User ID", "Total Present", "Total Late", "Total Hours", "Total WFH"]
            headers.extend([f"Day {d}" for d in range(1, days_in_month + 1)])

            header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="2E5090", end_color="2E5090", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            header_border = Border(
                left=Side(style="thin", color="FFFFFF"),
                right=Side(style="thin", color="FFFFFF"),
                top=Side(style="thin", color="FFFFFF"),
                bottom=Side(style="thin", color="FFFFFF"),
            )

            cell_border = Border(
                left=Side(style="thin", color="D0D0D0"),
                right=Side(style="thin", color="D0D0D0"),
                top=Side(style="thin", color="D0D0D0"),
                bottom=Side(style="thin", color="D0D0D0"),
            )
            cell_alignment = Alignment(horizontal="center", vertical="center")
            normal_font = Font(name="Segoe UI", size=10)

            white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            green_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
            red_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
            blue_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
            orange_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")

            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = header_border

            row_idx = 2
            for uid, u_info in self.users.items():
                name = u_info.get("display_name", uid)
                user_log = self.data.get(uid.lower(), {})

                present_cnt = 0
                late_cnt = 0
                wfh_cnt = 0
                total_hours = 0.0

                ws.cell(row=row_idx, column=1, value=name)
                ws.cell(row=row_idx, column=2, value=uid)

                for day in range(1, days_in_month + 1):
                    day_key = f"{day:02d}"
                    entry = user_log.get(day_key, {})
                    t_in = entry.get("in", "")
                    t_out = entry.get("out", "")
                    is_wfh = entry.get("wfh", False)

                    cell_val = ""
                    if t_in:
                        cell_val = f"In: {t_in}"
                        present_cnt += 1
                        try:
                            h, m = map(int, t_in.split(":"))
                            if h > self.late_hour or (h == self.late_hour and m > self.late_min):
                                late_cnt += 1
                        except (ValueError, AttributeError):
                            pass
                        if is_wfh:
                            wfh_cnt += 1
                    if t_out:
                        cell_val += f" / Out: {t_out}"
                        try:
                            d1 = datetime.strptime(t_in, "%H:%M")
                            d2 = datetime.strptime(t_out, "%H:%M")
                            total_hours += (d2 - d1).total_seconds() / 3600
                        except ValueError:
                            pass

                    col_idx = 6 + day
                    ws.cell(row=row_idx, column=col_idx, value=cell_val)

                ws.cell(row=row_idx, column=3, value=present_cnt)
                ws.cell(row=row_idx, column=4, value=late_cnt)
                ws.cell(row=row_idx, column=5, value=round(total_hours, 2))
                ws.cell(row=row_idx, column=6, value=wfh_cnt)
                row_idx += 1

            for row in range(2, row_idx):
                row_fill = white_fill if row % 2 == 0 else light_fill
                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.font = normal_font
                    cell.border = cell_border
                    cell.alignment = cell_alignment
                    if col == 3:
                        cell.fill = green_fill
                    elif col == 4:
                        cell.fill = red_fill
                    elif col == 5:
                        cell.fill = blue_fill
                    elif col == 6:
                        cell.fill = orange_fill
                    else:
                        cell.fill = row_fill

            for col in range(1, len(headers) + 1):
                column_letter = get_column_letter(col)
                if col == 1:
                    ws.column_dimensions[column_letter].width = 25
                elif col == 2:
                    ws.column_dimensions[column_letter].width = 15
                elif 3 <= col <= 6:
                    ws.column_dimensions[column_letter].width = 14
                else:
                    ws.column_dimensions[column_letter].width = 20

            ws.freeze_panes = "C2"
            wb.save(self.path)

            self.finished_export.emit(
                True,
                f"✅ Professional attendance report saved to:\n{self.path}\n\n"
                "✨ Features:\n• Bold headers with blue background\n"
                "• Color-coded stats (green/red/blue/orange)\n"
                "• Borders on all cells\n• Auto-sized columns",
            )
        except Exception as exc:
            self.finished_export.emit(False, str(exc))

    def stop(self):
        """Request the thread to stop at its next safe checkpoint."""
        self.requestInterruption()


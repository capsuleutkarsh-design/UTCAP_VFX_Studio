from PySide6.QtCore import QObject, Slot, Signal
import logging
import os

class WebBridge(QObject):
    """
    Bridge between Python and JavaScript for the Document Editor.
    Exposed to JS as 'pyBridge' via QWebChannel.
    """
    # Signals to send data TO JavaScript
    # msg_type: 'info', 'error', 'success'
    notification = Signal(str, str) 
    
    # content: base64/string, type: 'sheet' | 'doc'
    fileLoaded = Signal(str, str) 

    def __init__(self, parent=None):
        super().__init__(parent)
        self._logger = logging.getLogger("WebBridge")

    @Slot(str)
    def log(self, message):
        """Receives log messages from JavaScript."""
        self._logger.info(f"[JS] {message}")

    @Slot(str, str)
    def saveFile(self, filename, content):
        """
        Called by JS to save file content.
        """
        import base64
        
        try:
            self._logger.info(f"JS requested save: {filename} ({len(content)} chars)")
            
            if content.startswith("data:"):
                # Handle base64 encoded strings
                _, b64_data = content.split(",", 1)
                with open(filename, 'wb') as f:
                    f.write(base64.b64decode(b64_data))
            else:
                with open(filename, 'w', encoding='utf-8') as f:
                    f.write(content)
                
            self.notification.emit("success", f"File saved: {filename}")
        except Exception as e:
            self._logger.error(f"Failed to save file: {e}")
            self.notification.emit("error", f"Save failed: {str(e)}")

    @Slot(str)
    def loadFile(self, filename):
        """
        Called by JS (or Python wrapper) to read file content.
        Emits 'fileLoaded' signal with content.
        """
        try:
            content = ""
            file_type = 'doc'
            
            if filename.lower().endswith('.xlsx'):
                # Basic XLSX Importer
                import json
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(filename, data_only=True)
                    ws = wb.active
                    
                    # Construct Univer Data Structure
                    cell_data = {}
                    
                    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                        row_data = {}
                        has_data = False
                        for col_idx, value in enumerate(row):
                            if value is not None:
                                # Univer expects string values in 'v'
                                row_data[str(col_idx)] = {"v": str(value)}
                                has_data = True
                        
                        if has_data:
                            cell_data[str(row_idx)] = row_data
                            
                    univer_data = {
                        "id": "workbook-import",
                        "sheetOrder": ["sheet-01"],
                        "sheets": {
                            "sheet-01": {
                                "id": "sheet-01",
                                "name": ws.title,
                                "cellData": cell_data,
                                "rowCount": max(100, ws.max_row + 20),
                                "columnCount": max(20, ws.max_column + 5)
                            }
                        }
                    }
                    content = json.dumps(univer_data)
                    file_type = 'sheet'
                    
                except ImportError:
                    self._logger.error("openpyxl not found. Cannot import XLSX.")
                    self.notification.emit("error", "Missing 'openpyxl' library for Excel import.")
                    return
            else:
                # Text/JSON handling
                with open(filename, 'r', encoding='utf-8') as f:
                    content = f.read()
                file_type = 'sheet' if filename.endswith('.json') else 'doc'
            
            self.fileLoaded.emit(content, file_type)
            self.notification.emit("info", f"Loaded: {os.path.basename(filename)}")
        except Exception as e:
            self._logger.error(f"Failed to load file: {e}")
            self.notification.emit("error", f"Load failed: {str(e)}")

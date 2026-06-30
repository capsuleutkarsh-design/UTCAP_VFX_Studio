import logging
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from typing import List, Optional, Any
from ..models.shot_model import Shot, DepartmentInfo, FeedbackEntry, ArtistLogEntry
import os
from datetime import datetime
import shutil
from pathlib import Path
from ut_vfx.utils.security import SecurityValidator

class ExcelHandler:
    def __init__(self, filepath: str, project_config=None):
        self.filepath = filepath
        self.project_config = project_config
        self.workbook = None
        self.worksheet = None
        self._data_loaded = False
        self._row_map = None
    
    def _get_col_idx(self, field_name: str) -> int:
        if not self.project_config:
            return -1
        col_letter = self.project_config.column_mapping.get(field_name, "")
        if not col_letter:
            return -1
        return column_index_from_string(col_letter) - 1
    
    def load(self, data_only=False):
        # SECURITY CHECK
        try:
            is_valid, error = SecurityValidator.validate_excel_file(Path(self.filepath))
            if not is_valid:
                logging.error(f"Security Error loading {self.filepath}: {error}")
                return False
        except Exception as e:
            logging.exception(f"Security validation exception: {e}")
            return False

        if not os.path.exists(self.filepath):
            logging.info(f"File not found: {self.filepath}")
            return False
            
        import time
        retries = 3
        for attempt in range(retries):
            try:
                self.workbook = load_workbook(self.filepath, data_only=data_only)
                
                sheet_name = "MASTER"
                if self.project_config:
                    sheet_name = self.project_config.sheet_name
                
                if sheet_name in self.workbook.sheetnames:
                    self.worksheet = self.workbook[sheet_name]
                else:
                    self.worksheet = self.workbook.active
                self._row_map = None
                return True
            except Exception as e:
                if attempt < retries - 1:
                    logging.warning(f"File locked, retrying load ({attempt+1}/{retries}): {self.filepath}")
                    time.sleep(1.0)
                else:
                    logging.exception(f"Error loading workbook after {retries} attempts: {e}")
                    return False

    @staticmethod
    def _normalize_text(value: Any) -> str:
        return str(value).strip() if value is not None else ""

    def _build_row_map(self):
        if self._row_map is not None:
            return
        self._row_map = {}
        if not self.worksheet:
            return

        start_row = 3
        if self.project_config:
            start_row = self.project_config.data_start_row

        shot_col = self._get_col_idx("shot_name")
        if shot_col < 0:
            # Fallback to common AK74/standard column D
            shot_col = 3

        for row_idx, row in enumerate(
            self.worksheet.iter_rows(min_row=start_row, values_only=True), start=start_row
        ):
            if not row or shot_col >= len(row):
                continue
            shot_name = self._normalize_text(row[shot_col])
            if shot_name:
                self._row_map[shot_name.casefold()] = row_idx

    def _find_row_idx_by_shot_name(self, shot_name: str) -> int:
        self._build_row_map()
        if not self._row_map:
            return 0
        return self._row_map.get(self._normalize_text(shot_name).casefold(), 0)

    def _resolve_row_idx(self, shot: Shot) -> int:
        row_idx = int(getattr(shot, "_row_idx", 0) or 0)
        if row_idx > 0:
            return row_idx
        if not shot.shot_name:
            return 0
        row_idx = self._find_row_idx_by_shot_name(shot.shot_name)
        if row_idx > 0:
            shot._row_idx = row_idx
        return row_idx

    def _write_mapped_field(self, row_idx: int, field_names, value) -> bool:
        if not self.worksheet:
            return False
        if isinstance(field_names, str):
            field_names = [field_names]
        wrote = False
        for field_name in field_names:
            col_idx = self._get_col_idx(field_name)
            if col_idx >= 0:
                self.worksheet.cell(row=row_idx, column=col_idx + 1, value=value if value is not None else "")
                wrote = True
        return wrote

    @staticmethod
    def _latest_feedback_text(entries: List[FeedbackEntry]) -> str:
        if not entries:
            return ""
        latest = entries[-1]
        return str(getattr(latest, "text", "") or "")

    def _write_department(self, row_idx: int, dept_name: str, dept: DepartmentInfo):
        if dept_name == "comp":
            self._write_mapped_field(row_idx, ["comp_artist"], dept.artist)
            self._write_mapped_field(row_idx, ["comp_status"], dept.status)
            self._write_mapped_field(row_idx, ["comp_bid", "comp_mandays"], dept.bid_days)
            self._write_mapped_field(row_idx, ["comp_eta", "comp_target"], dept.target or dept.eta)
            return

        self._write_mapped_field(row_idx, [f"{dept_name}_artist"], dept.artist)
        self._write_mapped_field(
            row_idx,
            [f"{dept_name}_status", f"{dept_name}_required", f"{dept_name}_comp"],
            dept.status,
        )
        self._write_mapped_field(row_idx, [f"{dept_name}_bid", f"{dept_name}_mandays"], dept.bid_days)
        self._write_mapped_field(row_idx, [f"{dept_name}_eta", f"{dept_name}_target"], dept.target or dept.eta)

    def _write_shot_to_row(self, row_idx: int, shot: Shot):
        # Core shot-level fields
        self._write_mapped_field(row_idx, "reel", shot.reel_episode)
        self._write_mapped_field(row_idx, "shot_name", shot.shot_name)
        self._write_mapped_field(row_idx, "overall_status", shot.status)
        self._write_mapped_field(row_idx, "frames", shot.edit_frames)
        self._write_mapped_field(row_idx, "sow", shot.sow)
        self._write_mapped_field(row_idx, "target", shot.target)
        self._write_mapped_field(row_idx, ["assigned_artist", "artist_all"], shot.assigned_artist)
        self._write_mapped_field(row_idx, "version", shot.curr_version)
        self._write_mapped_field(row_idx, "latest_version", shot.curr_version)
        self._write_mapped_field(row_idx, "prev_version", shot.prev_version)
        self._write_mapped_field(row_idx, "scan_status", shot.scan_status)
        self._write_mapped_field(row_idx, "edit_status", shot.edit_status)
        self._write_mapped_field(row_idx, "in_os", shot.in_os)
        self._write_mapped_field(row_idx, "thumbnail", shot.thumbnail_path)
        self._write_mapped_field(row_idx, "wip_date", shot.wip_date)
        self._write_mapped_field(row_idx, "shot_done_date", shot.shot_done_date)
        self._write_mapped_field(row_idx, "submission_date", shot.submission_date)
        self._write_mapped_field(row_idx, "exr_date", shot.exr_submission)
        self._write_mapped_field(row_idx, "mov_date", shot.mov_submission)
        self._write_mapped_field(row_idx, "priority", shot.priority)
        self._write_mapped_field(row_idx, "shot_type", shot.shot_type)
        self._write_mapped_field(row_idx, "description", shot.description)
        self._write_mapped_field(row_idx, ["notes", "shot_comment"], shot.notes)

        # Feedback / comments
        self._write_mapped_field(
            row_idx,
            "internal_comment",
            shot.notes or self._latest_feedback_text(shot.feedback_internal),
        )
        self._write_mapped_field(row_idx, "client_feedback", self._latest_feedback_text(shot.feedback_client))
        self._write_mapped_field(row_idx, "director_feedback", self._latest_feedback_text(shot.feedback_director))
        latest_any = (
            self._latest_feedback_text(shot.feedback_internal)
            or self._latest_feedback_text(shot.feedback_client)
            or self._latest_feedback_text(shot.feedback_director)
        )
        self._write_mapped_field(row_idx, "latest_feedback", latest_any)

        # Department status mirrors
        self._write_mapped_field(row_idx, "internal_status", shot.comp_dept.status or shot.status)
        self._write_department(row_idx, "comp", shot.comp_dept)
        self._write_department(row_idx, "roto", shot.roto_dept)
        self._write_department(row_idx, "prep", shot.prep_dept)
        self._write_department(row_idx, "dmp", shot.dmp_dept)
        self._write_department(row_idx, "cg", shot.cg_dept)
        self._write_department(row_idx, "mgfx", shot.mgfx_dept)
        self._write_department(row_idx, "slapcomp", shot.slapcomp_dept)
    
    def read_shots(self) -> List[Shot]:
        if not self.load(data_only=True) or not self.worksheet:
            return []
        self._data_loaded = True
        
        shots = []
        start_row = 3
        if self.project_config:
            start_row = self.project_config.data_start_row
        
        for row_idx, row in enumerate(self.worksheet.iter_rows(min_row=start_row, values_only=True), start=start_row):
            shot = self._parse_row(row, row_idx)
            if shot:
                shots.append(shot)
        
        # Load extended data
        self._load_ut_data(shots)
        self._load_artist_log(shots)
        self._load_feedback_log(shots)
        
        return shots
    
    def _parse_row(self, row, row_idx: int) -> Optional[Shot]:
        def get_val(field_name: str, default=""):
            idx = self._get_col_idx(field_name)
            if idx < 0 or idx >= len(row):
                return default
            val = row[idx]
            return str(val) if val is not None else default
        
        def get_float(field_name: str, default=0.0):
            idx = self._get_col_idx(field_name)
            if idx < 0 or idx >= len(row):
                return default
            try:
                val = row[idx]
                return float(val) if val is not None else default
            except (TypeError, ValueError):
                return default
        
        shot_name = get_val("shot_name")
        if not shot_name:
            return None
        
        shot = Shot(
            shot_name=shot_name,
            reel_episode=get_val("reel"),
            status=get_val("overall_status", "WIP"),
            edit_frames=get_float("frames"),
            sow=get_val("sow"),
            assigned_artist=get_val("assigned_artist"),
            curr_version=get_val("version") or get_val("latest_version"),
            prev_version=get_val("prev_version"),
            target=get_val("target"),
            scan_status=get_val("scan_status"),
            edit_status=get_val("edit_status"),
            in_os=get_val("in_os"),
            thumbnail_path=get_val("thumbnail"),
            wip_date=get_val("wip_date"),
            shot_done_date=get_val("shot_done_date"),
            submission_date=get_val("submission_date"),
            exr_submission=get_val("exr_date"),
            mov_submission=get_val("mov_date"),
            
            comp_dept=DepartmentInfo(
                artist=get_val("assigned_artist"),
                status=get_val("internal_status") or get_val("overall_status"),
                eta=get_val("comp_eta")
            ),
            roto_dept=DepartmentInfo(
                artist=get_val("roto_artist"),
                status=get_val("roto_status"),
                bid_days=get_float("roto_bid"),
                eta=get_val("roto_eta")
            ),
            dmp_dept=DepartmentInfo(
                status=get_val("dmp_status"),
                bid_days=get_float("dmp_mandays"),
                eta=get_val("dmp_eta")
            ),
            cg_dept=DepartmentInfo(
                artist=get_val("cg_artist"),
                status=get_val("cg_status"),
                bid_days=get_float("cg_mandays"),
                eta=get_val("cg_eta")
            ),
            slapcomp_dept=DepartmentInfo(
                artist=get_val("slapcomp_artist"),
                status=get_val("slapcomp_status"),
                bid_days=get_float("slapcomp_bid"),
                eta=get_val("slapcomp_eta"),
                target=get_val("slapcomp_target")
            ),
            priority=int(get_float("priority", 3.0)),
            shot_type=get_val("shot_type")
        )
        
        # Parse feedback
        internal = get_val("internal_comment")
        if internal:
            shot.feedback_internal.append(FeedbackEntry(text=internal, source="Internal"))
        
        client = get_val("client_feedback")
        if client:
            shot.feedback_client.append(FeedbackEntry(text=client, source="Client"))
        
        director = get_val("director_feedback")
        if director:
            shot.feedback_director.append(FeedbackEntry(text=director, source="Director"))
        
        shot._row_idx = row_idx
        return shot
    
    def _load_ut_data(self, shots: List[Shot]):
        if not self.workbook or "UTVFX_DATA" not in self.workbook.sheetnames:
            return
        
        ws = self.workbook["UTVFX_DATA"]
        shot_map = {s.shot_name: s for s in shots}
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            shot_id = str(row[0])
            if shot_id in shot_map:
                shot = shot_map[shot_id]
                # Only overwrite if not already set from main sheet
                if not shot.shot_type:
                    shot.shot_type = str(row[1]) if len(row) > 1 and row[1] else ""
                if shot.priority == 3: # Default is 3
                    shot.priority = int(row[2]) if len(row) > 2 and row[2] else 3
                shot.is_hero = bool(row[3]) if len(row) > 3 else False
                shot.similar_to = str(row[4]).split(",") if len(row) > 4 and row[4] else []
    
    def _load_artist_log(self, shots: List[Shot]):
        if not self.workbook or "ARTIST_LOG" not in self.workbook.sheetnames:
            return
        
        ws = self.workbook["ARTIST_LOG"]
        shot_map = {s.shot_name: s for s in shots}
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            shot_id = str(row[0])
            if shot_id in shot_map:
                entry = ArtistLogEntry(
                    shot_id=shot_id,
                    artist=str(row[1]) if len(row) > 1 and row[1] else "",
                    department=str(row[2]) if len(row) > 2 and row[2] else "",
                    start_date=str(row[3]) if len(row) > 3 and row[3] else None,
                    end_date=str(row[4]) if len(row) > 4 and row[4] else None,
                    notes=str(row[5]) if len(row) > 5 and row[5] else ""
                )
                shot_map[shot_id].artist_history.append(entry)
    
    def _load_feedback_log(self, shots: List[Shot]):
        if not self.workbook or "FEEDBACK_LOG" not in self.workbook.sheetnames:
            return
        
        ws = self.workbook["FEEDBACK_LOG"]
        shot_map = {s.shot_name: s for s in shots}
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            shot_id = str(row[0])
            if shot_id in shot_map:
                entry = FeedbackEntry(
                    date=str(row[1]) if len(row) > 1 and row[1] else "",
                    source=str(row[2]) if len(row) > 2 and row[2] else "",
                    text=str(row[3]) if len(row) > 3 and row[3] else "",
                    logged_by=str(row[4]) if len(row) > 4 and row[4] else ""
                )
                source = entry.source.lower()
                if "client" in source:
                    shot_map[shot_id].feedback_client.append(entry)
                elif "director" in source:
                    shot_map[shot_id].feedback_director.append(entry)
                else:
                    shot_map[shot_id].feedback_internal.append(entry)
    
    def write_shots(self, shots: List[Shot]):
        if not self.load(data_only=False):
            return False

        self._build_row_map()
        for shot in shots:
            row_idx = self._resolve_row_idx(shot)
            if row_idx <= 0:
                continue
            self._write_shot_to_row(row_idx, shot)
        
        self._write_ut_data(shots)
        
        try:
            if self.workbook:
                import time
                retries = 3
                for attempt in range(retries):
                    try:
                        self.workbook.save(self.filepath)
                        logging.info(f"Saved {len(shots)} shots to {self.filepath}")
                        return True
                    except Exception as e:
                        if attempt < retries - 1:
                            logging.warning(f"File locked, retrying save ({attempt+1}/{retries}): {self.filepath}")
                            time.sleep(1.0)
                        else:
                            raise e
            return False
        except Exception as e:
            logging.exception(f"Error saving: {e}")
            return False

    def update_shot_field(self, shot_name: str, field: str, value, current_version: int = 0) -> bool:
        """
        Excel-mode compatibility with SQLiteHandler.update_shot_field API.
        current_version is ignored for Excel mode.
        """
        if not self.load(data_only=False):
            return False

        row_idx = self._find_row_idx_by_shot_name(shot_name)
        if row_idx <= 0:
            return False

        field_aliases = {
            "status": ["overall_status", "internal_status"],
            "assigned_artist": ["assigned_artist", "artist_all", "comp_artist"],
            "artist": ["assigned_artist", "artist_all", "comp_artist"],
            "curr_version": ["version", "latest_version"],
            "version": ["version", "latest_version"],
            "notes": ["notes", "shot_comment", "internal_comment"],
            "internal_comment": ["internal_comment", "shot_comment"],
            "client_feedback": ["client_feedback"],
            "director_feedback": ["director_feedback"],
            "target": ["target"],
            "scan_status": ["scan_status"],
            "edit_status": ["edit_status"],
            "in_os": ["in_os"],
        }

        target_fields = field_aliases.get(field, [field])
        if not self._write_mapped_field(row_idx, target_fields, value):
            return False

        try:
            if self.workbook:
                import time
                retries = 3
                for attempt in range(retries):
                    try:
                        self.workbook.save(self.filepath)
                        return True
                    except Exception as e:
                        if attempt < retries - 1:
                            logging.warning(f"File locked, retrying update save ({attempt+1}/{retries}): {self.filepath}")
                            time.sleep(1.0)
                        else:
                            raise e
            return False
        except Exception as e:
            logging.exception(f"Error saving field update for {shot_name}: {e}")
            return False
    
    def _write_ut_data(self, shots: List[Shot]):
        if not self.workbook:
            return
            
        if "UTVFX_DATA" not in self.workbook.sheetnames:
            self.workbook.create_sheet("UTVFX_DATA")
            ws = self.workbook["UTVFX_DATA"]
            ws.append(["SHOT_ID", "TYPE", "PRIORITY", "IS_HERO", "SIMILAR_TO", "CREATED", "UPDATED"])
        
        ws = self.workbook["UTVFX_DATA"]
        existing = {}
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row and row[0]:
                existing[str(row[0])] = row_idx
        
        now = datetime.now().strftime("%Y-%m-%d %H:%M")
        
        for shot in shots:
            similar_str = ",".join(shot.similar_to) if shot.similar_to else ""
            
            if shot.shot_name in existing:
                row_idx = existing[shot.shot_name]
                ws.cell(row=row_idx, column=2, value=shot.shot_type)
                ws.cell(row=row_idx, column=3, value=shot.priority)
                ws.cell(row=row_idx, column=4, value=shot.is_hero)
                ws.cell(row=row_idx, column=5, value=similar_str)
                ws.cell(row=row_idx, column=7, value=now)
            else:
                ws.append([shot.shot_name, shot.shot_type, shot.priority, shot.is_hero, similar_str, now, now])
    
    def create_backup(self):
        if not os.path.exists(self.filepath):
            return None
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_dir = os.path.join(os.path.dirname(self.filepath), "backups")
        os.makedirs(backup_dir, exist_ok=True)
        backup_path = os.path.join(backup_dir, f"backup_{timestamp}.xlsx")
        try:
            shutil.copy2(self.filepath, backup_path)
            logging.info(f"Backup created: {backup_path}")
            return backup_path
        except Exception as e:
            logging.error(f"Failed to create backup: {e}")
            return None

    def debug_column_mapping(self):
        """Prints the current column mapping for debugging."""
        if not self.project_config:
            logging.info("No project config loaded.")
            return
            
        logging.info(f"--- Debug Column Mapping for {self.project_config.code} ---")
        logging.info(f"Excel Path: {self.filepath}")
        logging.info(f"Sheet Name: {self.project_config.sheet_name}")
        
        # Load headers to verify
        if self.load(data_only=True) and self.worksheet:
            header_row = self.project_config.header_row
            headers = []
            for cell in self.worksheet[header_row]:
                headers.append(str(cell.value))
            logging.info(f"Headers found in row {header_row}: {headers}")
            
            logging.info("\nMapping:")
            for field, col_letter in self.project_config.column_mapping.items():
                idx = column_index_from_string(col_letter) - 1
                header_val = headers[idx] if 0 <= idx < len(headers) else "OUT OF BOUNDS"
                logging.info(f"  {field}: {col_letter} (Index {idx}) -> Header: '{header_val}'")
        else:
            logging.info("Could not load workbook to verify headers.")

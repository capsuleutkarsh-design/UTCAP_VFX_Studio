from PySide6.QtWidgets import (QDialog, QVBoxLayout, QFormLayout, QLineEdit, 
                             QDialogButtonBox, QPushButton, QFileDialog, QMessageBox, QHBoxLayout, 
                             QLabel, QComboBox, QGroupBox, QSpinBox)
from openpyxl import load_workbook
import os

class EditProjectDialog(QDialog):
    def __init__(self, project_config, parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"Edit Project: {project_config.code}")
        self.setMinimumWidth(600)
        self.project_config = project_config
        
        self.main_layout = QVBoxLayout(self)
        self.form = QFormLayout()
        
        # Code (Read Only)
        self.code_input = QLineEdit(project_config.code)
        self.code_input.setReadOnly(True)
        from ut_vfx.core.infra.design_tokens import ColorTokens as C
        self.code_input.setStyleSheet(f"color: {C.TEXT_GRAY_LIGHT}; background-color: {C.BG_HOVER};")
        self.form.addRow("Project Code:", self.code_input)
        
        # Name
        self.name_input = QLineEdit(project_config.name)
        self.form.addRow("Project Name:", self.name_input)
        
        # Excel Path
        self.excel_input = QLineEdit(project_config.excel_path)
        excel_layout = QHBoxLayout()
        excel_layout.addWidget(self.excel_input)
        self.excel_btn = QPushButton("Browse...")
        self.excel_btn.clicked.connect(self.browse_excel)
        excel_layout.addWidget(self.excel_btn)
        self.form.addRow("Excel File:", excel_layout)
        
        # Sheet Selection (Dynamic)
        self.sheet_combo = QComboBox()
        self.sheet_combo.setEditable(True)
        self.sheet_combo.setPlaceholderText("Select or Type Sheet Name")
        if project_config.sheet_name:
            self.sheet_combo.addItem(project_config.sheet_name)
            self.sheet_combo.setCurrentText(project_config.sheet_name)
            
        self.sheet_btn = QPushButton("Analyze File")
        self.sheet_btn.setToolTip("Re-scan Excel file for sheets")
        self.sheet_btn.clicked.connect(self.analyze_excel)
        
        sheet_layout = QHBoxLayout()
        sheet_layout.addWidget(self.sheet_combo)
        sheet_layout.addWidget(self.sheet_btn)
        self.form.addRow("Sheet Name:", sheet_layout)
        
        # Folder Base
        self.folder_input = QLineEdit(project_config.folder_base)
        folder_layout = QHBoxLayout()
        folder_layout.addWidget(self.folder_input)
        self.folder_btn = QPushButton("Browse...")
        self.folder_btn.clicked.connect(self.browse_folder)
        folder_layout.addWidget(self.folder_btn)
        self.form.addRow("Project Root:", folder_layout)
        
        self.main_layout.addLayout(self.form)
        
        # Advanced Config (Collapsible-ish via GroupBox)
        self.adv_group = QGroupBox("Advanced Layout Configuration")
        self.adv_group.setCheckable(True)
        self.adv_group.setChecked(False) # Collapsed by default
        adv_layout = QFormLayout()
        
        row_layout = QHBoxLayout()
        self.header_row_spin = QSpinBox()
        self.header_row_spin.setRange(1, 100); self.header_row_spin.setValue(project_config.header_row)
        row_layout.addWidget(QLabel("Header Row:"))
        row_layout.addWidget(self.header_row_spin)
        
        self.data_row_spin = QSpinBox()
        self.data_row_spin.setRange(1, 100); self.data_row_spin.setValue(project_config.data_start_row)
        row_layout.addWidget(QLabel("Data Start Row:"))
        row_layout.addWidget(self.data_row_spin)
        
        adv_layout.addRow("Excel Rows:", row_layout)
        self.adv_group.setLayout(adv_layout)
        
        self.main_layout.addWidget(self.adv_group)
        
        # Buttons
        self.buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        self.buttons.accepted.connect(self.validate_and_accept)
        self.buttons.rejected.connect(self.reject)
        self.main_layout.addWidget(self.buttons)
        
    def browse_excel(self):
        path, _ = QFileDialog.getOpenFileName(self, "Select Excel File", self.excel_input.text(), "Excel Files (*.xlsx *.xls)")
        if path:
            self.excel_input.setText(path)
            self.analyze_excel()
            
    def analyze_excel(self):
        path = self.excel_input.text().strip()
        if not path or not os.path.exists(path):
            return

        try:
            # Read-only for speed
            wb = load_workbook(path, read_only=True, keep_vba=False)
            sheets = wb.sheetnames
            wb.close()
            
            current = self.sheet_combo.currentText()
            self.sheet_combo.clear()
            self.sheet_combo.addItems(sheets)
            
            # If current is valid, keep it. Else try smart match.
            if current in sheets:
                 self.sheet_combo.setCurrentText(current)
            else:
                # Smart Selection
                best_match = None
                for sheet in sheets:
                    s_upper = sheet.upper()
                    if "MASTER" in s_upper:
                        best_match = sheet
                        break
                    if "TRACKER" in s_upper:
                        best_match = sheet
                
                if best_match:
                    self.sheet_combo.setCurrentText(best_match)
                
            QMessageBox.information(self, "Analysis Complete", f"Found {len(sheets)} sheets.")
            
        except Exception as e:
            QMessageBox.warning(self, "Analysis Failed", f"Could not read Excel file:\n{e}")
            
    def browse_folder(self):
        path = QFileDialog.getExistingDirectory(self, "Select Project Root Folder", self.folder_input.text())
        if path:
            self.folder_input.setText(path)
            
    def validate_and_accept(self):
        if not self.name_input.text():
            QMessageBox.warning(self, "Missing Data", "Project Name is required.")
            return
        self.accept()
        
    def get_data(self):
        return {
            "code": self.project_config.code,
            "name": self.name_input.text().strip(),
            "excel_path": self.excel_input.text().strip(),
            "folder_base": self.folder_input.text().strip(),
            "sheet_name": self.sheet_combo.currentText().strip(),
            "header_row": self.header_row_spin.value(),
            "data_start_row": self.data_row_spin.value()
        }

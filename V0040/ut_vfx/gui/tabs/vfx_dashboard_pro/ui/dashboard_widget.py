from PySide6.QtWidgets import (QWidget, QLabel, QLineEdit, QMessageBox, QMenu, QFileDialog, QInputDialog)
from PySide6.QtGui import QAction, QPixmap
from ut_vfx.core.infra.qt_compat import Qt, QTimer, Signal, QItemSelectionModel
from openpyxl import Workbook, load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.styles import Protection
from openpyxl.worksheet.datavalidation import DataValidation
from ..core.project_manager import ProjectManager
from ..core.excel_handler import ExcelHandler
from ..core.sqlite_handler import SQLiteHandler
from ..core.file_lock import FileLock
from ut_vfx.core.infra.app_context import AppContext
from ut_vfx.core.infra.database_manager import database_manager
from ut_vfx.core.infra.global_config import GlobalConfig
from ut_vfx.core.infra.theme_manager import ThemeManager
from ut_vfx.core.system.adaptation_engine import system_engine


# ... imports ...
from .shot_detail import ShotDetailWidget
from ..utils.thumbnail import ThumbnailGenerator
from ut_vfx.utils.async_image_loader import AsyncImageLoader
import os
import logging
from .add_project_dialog import AddProjectDialog
from .edit_project_dialog import EditProjectDialog
from datetime import datetime
from ..core.sqlite_handler import StaleDataError # Import Exception

from .history_dialog import HistoryDialog # Import UI
from ..core.poll_worker import PollWorker # Real-time updates
from .dashboard_layout_builder import build_dashboard_ui
from .dashboard_sync_service import DashboardSyncService
from .dashboard_avatar_service import DashboardAvatarService
from .sync_dialog import SyncDialog
from collections import deque
from ut_vfx.core.dcc_launcher import DCCLauncher
import shutil
import glob
from pathlib import Path
from PySide6.QtCore import QThread, Signal

class AutoPublishWorker(QThread):
    finished_signal = Signal(bool, str)
    
    def __init__(self, shot, project_manager, project_code):
        super().__init__()
        self.shot = shot
        self.project_manager = project_manager
        self.project_code = project_code
        
    def run(self):
        try:
            # 1. Find the comp output folder
            comp_path = self.project_manager.get_folder_path(
                self.project_code, "comp", self.shot.reel_episode, self.shot.shot_name
            )
            if not comp_path:
                self.finished_signal.emit(False, "Could not resolve Comp path.")
                return
                
            # Usually render is in 07_Comp/Output
            comp_output = Path(comp_path) / "Output"
            if not comp_output.exists():
                self.finished_signal.emit(False, "Comp Output folder not found.")
                return
                
            # 2. Find the output destination
            final_exr_path = self.project_manager.get_folder_path(
                self.project_code, "output", self.shot.reel_episode, self.shot.shot_name
            )
            if not final_exr_path:
                self.finished_signal.emit(False, "Could not resolve Final Output path.")
                return
                
            dest_exr = Path(final_exr_path) / "EXR"
            dest_mov = Path(final_exr_path) / "MOV"
            dest_exr.mkdir(parents=True, exist_ok=True)
            dest_mov.mkdir(parents=True, exist_ok=True)
            
            moved = 0
            # 3. Copy files over
            for file_path in comp_output.rglob("*.*"):
                if file_path.is_file():
                    ext = file_path.suffix.lower()
                    target_dir = dest_mov if ext in ['.mov', '.mp4'] else dest_exr
                    shutil.copy2(str(file_path), str(target_dir / file_path.name))
                    moved += 1
                    
            self.finished_signal.emit(True, f"Auto-published {moved} files to 08_Output.")
        except Exception as e:
            self.finished_signal.emit(False, str(e))


class ClickableLabel(QLabel):
    clicked = Signal()
    def mousePressEvent(self, event):
        self.clicked.emit()
        super().mousePressEvent(event)


from ..controllers.thumbnail_mixin import DashboardThumbnailMixin
from ..controllers.kanban_mixin import DashboardKanbanMixin
from ..controllers.filter_mixin import DashboardFilterMixin
from .components.dashboard_builder_mixin import DashboardBuilderMixin
from .components.dashboard_actions_mixin import DashboardActionsMixin
from .components.dashboard_project_mixin import DashboardProjectMixin

class DashboardWidget(
    DashboardBuilderMixin,
    DashboardActionsMixin,
    DashboardProjectMixin,
    DashboardFilterMixin, 
    DashboardThumbnailMixin, 
    DashboardKanbanMixin, 
    QWidget
):
    def _notify(self, message: str, level: str = "info", duration: int = 4000, details: str = ""):
        """Use main window feedback style when available, else fallback locally."""
        host = self.window()
        if host and hasattr(host, "show_feedback"):
            try:
                host.show_feedback(message=message, level=level, duration=duration, details=details)
                return
            except Exception:
                pass
        self.status_bar.showMessage(message, duration)
        if level == "error" and details:
            QMessageBox.critical(self, "Error", details)

    def _update_empty_state(self):
        if not hasattr(self, "view_stack") or not hasattr(self, "empty_state_frame"):
            return

        if not self.current_project:
            self.empty_state_title.setText("No project selected")
            self.empty_state_body.setText("Select a project from the dropdown to load shots and start planning.")
            self.view_stack.setCurrentWidget(self.empty_state_frame)
            return

        if self._is_artist_scope() and not self.all_shots:
            self.empty_state_title.setText("No shots assigned to you")
            self.empty_state_body.setText("Assigned shots will appear here automatically when production updates assignments.")
            self.view_stack.setCurrentWidget(self.empty_state_frame)
            return

        if self.local_mode and not self.all_shots:
            self.empty_state_title.setText("LOCAL MODE: no synced shots yet")
            self.empty_state_body.setText(
                "Import from Excel or connect to central database. Team sync actions are limited in LOCAL MODE."
            )
            self.view_stack.setCurrentWidget(self.empty_state_frame)
            return

        if self.all_shots and not self.displayed_shots:
            self.empty_state_title.setText("No matching shots")
            self.empty_state_body.setText("Try clearing search/filter selections to reveal available shots.")
            self.view_stack.setCurrentWidget(self.empty_state_frame)
            return

        board_mode = bool(getattr(self, "view_toggle_btn", None) and self.view_toggle_btn.isChecked())
        active_widget = self.kanban_board if board_mode else self.table
        self.view_stack.setCurrentWidget(active_widget)

    @staticmethod
    def _normalize_roles(roles_data):
        """Normalize role payload to lowercase canonical role names."""
        if isinstance(roles_data, str):
            raw_roles = [roles_data]
        elif isinstance(roles_data, list):
            raw_roles = roles_data
        else:
            raw_roles = ["Artist"]

        aliases = {
            "producer": "supervisor",
            "production": "supervisor",
            "pro": "supervisor",
            "coordinator": "supervisor",
            "coord": "supervisor",
            "dev": "developer",
        }

        normalized = []
        for role in raw_roles:
            role_text = str(role or "Artist").strip().lower()
            normalized.append(aliases.get(role_text, role_text))

        return normalized or ["artist"]

    @staticmethod
    def _is_local_fallback_mode() -> bool:
        try:
            status = database_manager.get_runtime_status() or {}
            mode = str(status.get("active_mode", "")).lower()
            return mode == "sqlite" and bool(status.get("fallback_used", False))
        except Exception:
            return False

    def _user_can_edit(self):
        return any(r in {"supervisor", "developer", "admin"} for r in self.user_roles)

    def _is_artist_scope(self) -> bool:
        """Artists should only see their assigned shots."""
        return "artist" in self.user_roles and not self._user_can_edit()

    def _artist_identity_candidates(self):
        candidates = {
            str(self.user_data.get("user_id", "")).strip(),
            str(self.user_data.get("username", "")).strip(),
            str(self.user_data.get("display_name", "")).strip(),
            str(self.user_display_name or "").strip(),
        }
        return {c.lower() for c in candidates if c}

    def _filter_shots_for_current_user(self, shots):
        if not self._is_artist_scope():
            return list(shots or [])

        identities = self._artist_identity_candidates()
        filtered = []
        for shot in shots or []:
            assigned = str(getattr(shot, "assigned_artist", "") or "").strip().lower()
            if assigned and assigned in identities:
                filtered.append(shot)
        return filtered



    def _warn_if_exr_policy_limits_project(self):
        """Warn once per project when EXR assets exist but EXR loading is disabled."""
        try:
            if GlobalConfig.exr_loading_enabled():
                return
            has_exr = False
            for shot in self.all_shots or []:
                scan_path = str(getattr(shot, "scan_path", "") or "").lower()
                render_path = str(getattr(shot, "render_path", "") or "").lower()
                if ".exr" in scan_path or ".exr" in render_path:
                    has_exr = True
                    break
            if has_exr:
                self.status_bar.showMessage(
                    "EXR assets detected. EXR loading is OFF (safe default). "
                    "Set UTVFX_ENABLE_EXR_LOADING=1 or enable_exr_loading=true.",
                    12000,
                )
        except Exception as exc:
            logging.debug("EXR policy warning check skipped: %s", exc)

    def showEvent(self, event):
        super().showEvent(event)
        if self.current_project and self.displayed_shots:
            self._queue_visible_thumbnails()
            if self._thumb_prefetch_queue and not self._thumb_prefetch_timer.isActive():
                self._thumb_prefetch_timer.start()

    def hideEvent(self, event):
        if self._thumb_prefetch_timer.isActive():
            self._thumb_prefetch_timer.stop()
        if self._visible_thumb_timer.isActive():
            self._visible_thumb_timer.stop()
        super().hideEvent(event)

    def log(self, message):
        """Log message using proper logging module."""
        try:
            logging.info(message)
        except Exception as e:
            logging.exception(f"Failed to log message: {e}")

    def _get_user_list(self):
        """Flatten user dict to list of display names. Filters by Role (Artist/Supervisor)."""
        valid_users = set()
        try:
            db_users = self.user_manager.get_all_users()
            for username, u in db_users.items():
                role = u.get('role')
                if not role:
                    roles = u.get('roles', [])
                    role = roles[0] if isinstance(roles, list) and roles else "Artist"
                name = u.get('display_name', '').strip()
                
                # Check normalized role (case-insensitive)
                # Allow any role that contains 'artist' or is in the specific list
                r_norm = str(role).lower()
                if (r_norm in ['artist', 'supervisor', 'lead', 'generalist', 'admin', 'producer'] or 
                    'artist' in r_norm):
                    if name:
                        valid_users.add(name)
                    else:
                        valid_users.add(username)
        except Exception as e:
            logging.exception(f"ERROR: Failed to fetch DB users: {e}")
        result = sorted(list(valid_users))
        if not result:
            result = ["Artist"]
        return result

    def init_ui(self):
        build_dashboard_ui(self)

    def _cleanup_poll_worker(self, timeout_ms: int = 2000):
        worker = self.poll_worker
        if worker is None:
            return
        try:
            if hasattr(worker, "stop"):
                try:
                    worker.stop(timeout_ms=timeout_ms)
                except TypeError:
                    worker.stop()
            elif worker.isRunning():
                worker.requestInterruption()
                worker.wait(timeout_ms)
        except Exception as exc:
            logging.debug("Poll worker shutdown warning: %s", exc)
        try:
            worker.deleteLater()
        except RuntimeError as exc:
            logging.debug("Poll worker deleteLater skipped: %s", exc)
        if self.poll_worker is worker:
            self.poll_worker = None

    def _cleanup_avatar_upload_worker(self, timeout_ms: int = 1500):
        worker = self.avatar_upload_worker
        if worker is None:
            return
        try:
            if worker.isRunning():
                if hasattr(worker, "stop"):
                    worker.stop()
                else:
                    worker.requestInterruption()
                worker.wait(timeout_ms)
        except Exception as exc:
            logging.debug("Avatar upload worker shutdown warning: %s", exc)
        try:
            worker.deleteLater()
        except RuntimeError as exc:
            logging.debug("Avatar worker deleteLater skipped: %s", exc)
        if self.avatar_upload_worker is worker:
            self.avatar_upload_worker = None

    def check_for_updates(self):
        if self._is_closing:
            return
        if not self.current_project:
            return

        if self.current_excel_path and os.path.exists(self.current_excel_path):
            try:
                current_mtime = os.path.getmtime(self.current_excel_path)
                if self.last_excel_mtime is None:
                    self.last_excel_mtime = current_mtime
                elif current_mtime > (self.last_excel_mtime + 0.5):
                    self.last_excel_mtime = current_mtime
                    self.log(f"Detected external Excel update: {self.current_excel_path}")
                    if isinstance(self.data_handler, SQLiteHandler):
                        synced = self._sync_excel_to_database(self.current_project.code, self.current_project)
                        if synced:
                            self.status_bar.showMessage(f"Imported {synced} updated shots from Excel")
                    self.refresh_data()
            except Exception as e:
                logging.debug(f"Excel update check failed: {e}")

    def refresh_data(self):
        if self._is_closing:
            return
        if self.current_project:
            logging.debug("Refreshing data...")
            self.switch_project(self.current_project.code)
            
    def open_query_builder(self):
        """Opens the Advanced Query Builder dialog."""
        from .query_builder_dialog import QueryBuilderDialog
        dialog = QueryBuilderDialog(self)
        
        # Populate with existing rules if any
        if self.advanced_query_rules:
            # Clear default
            for w in list(dialog.rules):
                dialog.remove_rule(w)
            for rule in self.advanced_query_rules:
                dialog.add_rule()
                w = dialog.rules[-1]
                w.field_combo.setCurrentText(rule['field'])
                w.op_combo.setCurrentText(rule['operator'])
                w.value_input.setText(rule['value'])
            idx = 0 if self.advanced_query_match_type == "AND" else 1
            dialog.match_combo.setCurrentIndex(idx)
            
        dialog.query_applied.connect(self._on_query_applied)
        dialog.exec()
        
    def _on_query_applied(self, rules, match_type):
        self.advanced_query_rules = rules
        self.advanced_query_match_type = match_type
        
        # Highlight button if active
        if rules:
            self.advanced_query_btn.setStyleSheet("background-color: #00E5FF; color: black; font-weight: bold;")
        else:
            self.advanced_query_btn.setStyleSheet("")
            
        self.apply_filters()
            
    def load_projects(self):
        projects = self.project_manager.get_all_projects()
        self.project_combo.clear()
        self.project_combo.addItem("Select Project...", None)
        for p in projects:
            self.project_combo.addItem(f"{p.code} - {p.name}", p.code)
            
        if self.project_manager.default_project:
            index = self.project_combo.findData(self.project_manager.default_project)
            if index >= 0:
                self.project_combo.setCurrentIndex(index)
                
    def on_project_changed(self, index):
        project_code = self.project_combo.currentData()
        if not project_code:
            self.current_project = None
            self.all_shots = []
            self.displayed_shots = []
            self.update_table()
            return
        self.switch_project(project_code)

    def update_table(self):
        self.table_model.update_data(self.displayed_shots)
        self.header_view.update_filters(self.displayed_shots)
        self.stats_widget.update_stats(self.displayed_shots)
        self.update_kanban()
        self._update_empty_state()
        
        # Update Users List
        self.users_list.populate(self.all_users)





    def start_thumbnail_loading(self):
        if not self.current_project:
            return
        self._cancel_thumbnail_prefetch()
        self._queue_visible_thumbnails()

        visible_names = {s.shot_name for s in self._visible_shots() if getattr(s, "shot_name", None)}
        for shot in self.displayed_shots:
            shot_name = getattr(shot, "shot_name", "")
            if not shot_name or shot_name in visible_names:
                continue
            self._enqueue_thumbnail_prefetch(shot)

        if self._thumb_prefetch_queue and self.isVisible():
            self._thumb_prefetch_timer.start()
                
    def cycle_theme(self):
        """Cycle through available themes."""
        try:
            if hasattr(ThemeManager, "get_available_themes"):
                themes = ThemeManager.get_available_themes()
                if not themes:
                    return
                current = getattr(self, "current_theme_idx", 0)
                current = (current + 1) % len(themes)
                self.current_theme_idx = current
                new_theme = themes[current]
                ThemeManager.apply_theme(new_theme)
            else:
                ThemeManager.toggle_mode()
                new_theme = "Dark" if ThemeManager.is_dark_mode() else "Light"

            if str(new_theme).lower() in {"default", "dark", "oceanic", "neon"}:
                self.theme_btn.setText("🌙")
            else:
                self.theme_btn.setText("☀")
        except Exception as e:
            logging.exception(f"Theme toggle failed: {e}")

    def on_image_loaded(self, identifier, path, image):
        self._thumb_requests_inflight.discard(identifier)
        identifier_project = self._project_code_from_identifier(identifier)
        if self.current_project and identifier_project and identifier_project != self.current_project.code:
            return
        pixmap = QPixmap.fromImage(image)
        if pixmap.isNull():
            return

        self.image_cache[identifier] = pixmap
        self.image_cache[path] = pixmap

        shot_name = self._shot_name_from_identifier(identifier)
        # Update Shot object
        # Find shot by name (identifier)
        for shot in self.all_shots:
            if shot.shot_name == shot_name:
                shot.thumbnail_path = path
                break
        
        # Force table update to replace Yellow with Result
        self.table.viewport().update()

        # If Detail Widget is open for this shot, update it
        if self.detail_widget and self.detail_widget.shot.shot_name == shot_name:
            self.detail_widget.thumb_label.setPixmap(pixmap.scaled(100, 56, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation))

    def on_image_started(self, identifier):
        """Called when loader starts processing a thumbnail (Show Yellow)."""
        identifier_project = self._project_code_from_identifier(identifier)
        if self.current_project and identifier_project and identifier_project != self.current_project.code:
            return
        yellow_path = self.thumb_gen.get_yellow_placeholder()
        shot_name = self._shot_name_from_identifier(identifier)

        # Update Shot object to point to yellow placeholder temporarily
        for shot in self.all_shots:
            if shot.shot_name == shot_name:
                shot.thumbnail_path = yellow_path
                break
        
        # Trigger table repaint
        self.table.viewport().update()

        # Update Detail Widget if open
        if self.detail_widget and self.detail_widget.shot.shot_name == shot_name:
            pix = QPixmap(yellow_path)
            self.detail_widget.thumb_label.setPixmap(pix.scaled(100, 56, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation))

    def on_item_clicked(self, index):
        """Handle single click to open detail dock (User Preference)."""
        if not index.isValid():
            return
            
        row = index.row()
        if 0 <= row < len(self.displayed_shots):
            shot = self.displayed_shots[row]
            # Debounce: If already open for this shot, do nothing? 
            # Actually, re-opening is fine, but maybe check if same?
            if self.detail_widget and self.detail_widget.shot.shot_name == shot.shot_name:
                if not self.detail_container.isVisible():
                    self.detail_container.show()
                return

            self.open_detail_dock(shot)

    def on_item_double_clicked(self, index):
        if not index.isValid(): return
        shot = self.displayed_shots[index.row()]
        self.open_detail_dock(shot)
        
    def open_detail_dock(self, shot):
        logging.debug(f"Opening detail dock for {shot.shot_name}")
        if self.detail_widget:
            self.detail_layout.removeWidget(self.detail_widget)
            self.detail_widget.deleteLater()

        # Ensure thumb refresh also covers stale/missing placeholder paths.
        if self._needs_thumbnail_refresh(shot):
            self._queue_thumbnail_load(shot)

        # Pass all_users to ShotDetailWidget
        self.detail_widget = ShotDetailWidget(
            shot,
            self.user_roles,
            self.project_manager,
            self.all_shots,
            self.all_users,
            user_data=self.user_data,
            current_project_code=(self.current_project.code if self.current_project else ""),
            inherit_app_theme=self.inherit_app_theme,
        )
            
        self.detail_widget.close_requested.connect(self.close_detail_dock)
        self.detail_widget.search_requested.connect(self.on_detail_search)
        self.detail_widget.save_requested.connect(self.on_shot_save) # Connect save signal
        self.detail_layout.addWidget(self.detail_widget)
        self.detail_container.show()
        
        # Ensure splitter gives it space
        if self.detail_container.width() < 50:
             sizes = self.splitter.sizes()
             # Allocate a larger default width for readability in SOW/feedback fields.
             if len(sizes) >= 2:
                 new_sizes = list(sizes)
                 total = max(sum(sizes), 1)
                 target_detail = max(system_engine.scale_px(460, minimum=380), int(total * 0.36))
                 available_for_others = max(100, total - target_detail)
                 new_sizes[-1] = target_detail
                 if len(new_sizes) == 2:
                     new_sizes[0] = available_for_others
                 else:
                     middle_width = new_sizes[1] if self.users_list.isVisible() else 0
                     new_sizes[1] = middle_width
                     new_sizes[0] = max(100, available_for_others - middle_width)
                 self.splitter.setSizes(new_sizes)

    def show_context_menu(self, pos):
        try:
            index = self.table.indexAt(pos)
            if not index.isValid():
                return
                
            row = index.row()
            if row < 0 or row >= len(self.displayed_shots):
                return
                
            shot = self.displayed_shots[row]
            
            menu = QMenu(self)
            
            # Quick Folders
            folders = {
                "Open Scan": "scan",
                "Open Roto": "roto",
                "Open Prep": "prep",
                "Open DMP": "dmp",
                "Open Comp": "comp",
                "Open Output": "output"
            }
            
            for label, key in folders.items():
                action = QAction(label, menu)
                # Use default argument to capture loop variable safely
                action.triggered.connect(lambda checked=False, k=key, s=shot: self.open_shot_folder(k, s))
                menu.addAction(action)
                
            menu.addSeparator()
            
            # DCC Launchers
            dcc_menu = menu.addMenu("Launch DCC")
            dcc_launcher = DCCLauncher(self)
            
            nuke_action = QAction("Foundry Nuke", dcc_menu)
            nuke_action.triggered.connect(lambda: dcc_launcher.launch("nuke", shot.id))
            dcc_menu.addAction(nuke_action)
            
            natron_action = QAction("Natron", dcc_menu)
            natron_action.triggered.connect(lambda: dcc_launcher.launch("natron", shot.id))
            dcc_menu.addAction(natron_action)
            
            blender_action = QAction("Blender", dcc_menu)
            blender_action.triggered.connect(lambda: dcc_launcher.launch("blender", shot.id))
            dcc_menu.addAction(blender_action)
            
            silhouette_action = QAction("BorisFX Silhouette", dcc_menu)
            silhouette_action.triggered.connect(lambda: dcc_launcher.launch("silhouette", shot.id))
            dcc_menu.addAction(silhouette_action)
            
            menu.addSeparator()
            
            # PHASE 3: View History
            history_action = QAction("View History", menu)
            history_action.triggered.connect(lambda: self.show_history_dialog(shot))
            menu.addAction(history_action)
                
            menu.exec(self.table.viewport().mapToGlobal(pos))
        except Exception as e:
            logging.exception(f"Context Menu Error: {e}")
            
    def open_shot_folder(self, folder_key, shot):
        if not self.current_project:
            return
            
        try:
            path = self.project_manager.get_folder_path(
                self.current_project.code, 
                folder_key, 
                shot.reel_episode, # Use reel_episode as per integrated app model
                shot.shot_name
            )
            
            if path:
                if not self.project_manager.open_folder(path):
                    self._notify(
                        "Could not open folder. Check if it exists.",
                        "warning",
                        details=path,
                    )
            else:
                self._notify(f"No path configured for '{folder_key}'.", "warning")
        except Exception as e:
            self._notify("Failed to open folder.", "error", details=str(e))

    def show_history_dialog(self, shot):
        if not self.current_project:
            return
            
        dialog = HistoryDialog(self.current_project.code, shot.shot_name, self)
        dialog.exec()
        
    def on_shot_save(self, shot):
        """Handle save request from Detail Widget with Optimistic Locking"""
        if not self.data_handler:
            return
        if not self._user_can_edit():
            self._notify("Read-only: only Supervisor/Developer/Admin can save changes.", "warning")
            return
            
        try:
            success = self.data_handler.write_shots([shot])
            if success:
                mirrored = self._mirror_shots_to_excel([shot])
                self.refresh_data()
                if mirrored:
                    self._notify(f"Shot {shot.shot_name} saved successfully.", "success")
                else:
                    self._notify(
                        f"Shot {shot.shot_name} saved to database, but Excel mirror failed.",
                        "warning",
                    )
                    
                # TRIGGER AUTO-PUBLISH IF APPROVED
                status = getattr(shot, 'status', '').upper()
                if status == 'APPROVED':
                    self._notify("Shot Approved! Auto-publishing renders to 08_Output...", "info")
                    self.publish_worker = AutoPublishWorker(shot, self.project_manager, self.current_project.code)
                    def on_publish_finished(ok, msg):
                        if ok:
                            self._notify(msg, "success")
                        else:
                            self._notify(f"Auto-publish failed: {msg}", "warning")
                    self.publish_worker.finished_signal.connect(on_publish_finished)
                    self.publish_worker.start()
                    
            else:
                self._notify("Could not save shot. Check logs.", "warning")
        except StaleDataError as e:
            reply = QMessageBox.warning(
                self, 
                "Conflict Detected", 
                f"{str(e)}\n\nAnother user has modified this shot since you opened it.\n"
                "You must reload the project to get the latest data before making changes.",
                QMessageBox.StandardButton.Refresh | QMessageBox.StandardButton.Cancel
            )
            if reply == QMessageBox.StandardButton.Refresh:
                self.refresh_data()
        except PermissionError as e:
            self._notify(str(e), "warning")
        except Exception as e:
            self._notify("An unexpected error occurred.", "error", details=str(e))

    def on_detail_search(self, text):
        # If text is empty, maybe use the shot name? Or just focus search bar
        if not text:
            self.search_input.setFocus()
            self.search_input.selectAll()
        else:
            self.search_input.setText(text)
            
    def change_avatar(self):
        """Open file picker and save new avatar."""
        if self._is_closing:
            return
        if not hasattr(self, "avatar_label"):
            self._notify("Profile avatar is managed in the main header.", "info")
            return

        path = self.avatar_service.choose_avatar_file(self)
        if not path:
            return

        try:
            self._cleanup_avatar_upload_worker(timeout_ms=1000)
            username = self.user_data.get("username", "unknown")
            self.avatar_upload_worker = self.avatar_service.start_avatar_upload(
                path,
                username,
                self.on_avatar_upload_finished,
            )
            if self.avatar_upload_worker:
                self._notify("Uploading profile picture in background...", "info")
            else:
                self._notify("No avatar file selected.", "warning")
        except Exception as e:
            self._notify("Failed to initialize avatar upload.", "error", details=str(e))

    def on_avatar_upload_finished(self, success, result_path, username):
        sender = self.sender()
        if sender is not None and sender is not self.avatar_upload_worker:
            return
        if self._is_closing:
            self._cleanup_avatar_upload_worker(timeout_ms=500)
            return
        if not success:
            self._notify("Could not copy avatar image.", "error", details=str(result_path))
            self.avatar_upload_worker = None
            return

        ok, payload = self.avatar_service.finalize_avatar_upload(success, result_path, username)
        if ok:
            self.load_user_avatar_from_path(payload)
        else:
            self._notify("Avatar update could not be saved to database.", "warning", details=str(payload))

        self.avatar_upload_worker = None

    def load_user_avatar(self):
        """Load avatar from DB on startup."""
        try:
            username = self.user_data.get('username', 'unknown')
            path = self.avatar_service.get_user_avatar_path(username)
            if path and os.path.exists(path):
                self.load_user_avatar_from_path(path)
        except Exception as e:
            logging.exception(f"Avatar load error: {e}")

    def load_user_avatar_from_path(self, path):
        """Helper to render the avatar."""
        if not hasattr(self, "avatar_label"):
            return

        self.avatar_service.apply_avatar_to_label(self.avatar_label, path, size=40)
        
    def close_detail_dock(self):
        self.detail_container.hide()
        if self.detail_widget:
            self.detail_widget.deleteLater()
            self.detail_widget = None

    def set_project_root_click(self):
        code = self.project_combo.currentData()
        if not code:
            self._notify("Please select a project first.", "warning")
            return
        current_root = ""
        if self.current_project:
            current_root = self.current_project.folder_base
            
        path = QFileDialog.getExistingDirectory(self, f"Select Root for {code}", current_root)
        if path:
            self.project_manager.set_project_folder_base(code, path)
            self._notify(f"Project root set to: {path}", "success")
            # Refresh to Apply
            self.switch_project(code)

    @staticmethod
    def _friendly_header_name(field_name: str) -> str:
        text = str(field_name or "").replace("_", " ").strip()
        return text.title() if text else ""

    def add_project_click(self):
        dialog = AddProjectDialog(self)
        if dialog.exec():
            data = dialog.get_data()
            try:
                self.project_manager.add_project(
                    code=data['code'],
                    name=data['name'],
                    excel_path=data['excel_path'],
                    folder_base=data['folder_base'],
                    sheet_name=data['sheet_name'],
                    header_row=data['header_row'],
                    data_start_row=data['data_start_row']
                )
                self.load_projects() # Refresh list
                # Select the new project
                idx = self.project_combo.findData(data['code'])
                if idx >= 0:
                    self.project_combo.setCurrentIndex(idx)
            except Exception as e:
                self._notify("Failed to create project.", "error", details=str(e))

    def delete_project_click(self):
        if not self.current_project:
            return
            
        code = self.current_project.code
        text, ok = QInputDialog.getText(self, "Delete Project", 
                                        f"WARNING: This will delete ALL data for '{code}'.\n\nType 'DELETE' to confirm:",
                                        QLineEdit.EchoMode.Normal, "")
        if ok and text == "DELETE":
            if self.project_manager.delete_project(code):
                self._notify(f"Project {code} deleted.", "success")
                self.load_projects()
                # Clear selection or select another
                if self.project_combo.count() > 0:
                    self.project_combo.setCurrentIndex(0)
            else:
                self._notify(f"Failed to delete project {code}.", "error")

    def show_column_menu(self):
        menu = QMenu(self)
        header = self.table.horizontalHeader()
        
        select_all = QAction("Select All", menu)
        select_all.triggered.connect(lambda: self._set_all_columns(True))
        menu.addAction(select_all)
        
        deselect_all = QAction("Deselect All", menu)
        deselect_all.triggered.connect(lambda: self._set_all_columns(False))
        menu.addAction(deselect_all)
        
        menu.addSeparator()
        
        for i in range(self.table_model.columnCount()):
            col_name = self.table_model.headerData(i, Qt.Orientation.Horizontal)
            action = QAction(col_name, menu)
            action.setCheckable(True)
            action.setChecked(not header.isSectionHidden(i))
            action.setData(i)
            action.triggered.connect(self.toggle_column)
            menu.addAction(action)
            
        menu.exec(self.columns_btn.mapToGlobal(self.columns_btn.rect().bottomLeft()))
        
    def toggle_column(self):
        action = self.sender()
        col_idx = action.data()
        if action.isChecked():
            self.table.showColumn(col_idx)
        else:
            self.table.hideColumn(col_idx)
            
    def _set_all_columns(self, visible: bool):
        for i in range(self.table_model.columnCount()):
            if visible:
                self.table.showColumn(i)
            else:
                self.table.hideColumn(i)

    def _is_excel_newer_than_db(self, project_code, excel_path):
        return self.sync_service.is_excel_newer_than_db(project_code, excel_path)

    def _sync_excel_to_database(self, project_code, project):
        return self.sync_service.sync_excel_to_database(project_code, project)

    def _mirror_shots_to_excel(self, shots):
        success, last_mtime = self.sync_service.mirror_shots_to_excel(
            shots=shots,
            current_project=self.current_project,
            data_handler=self.data_handler,
        )
        if last_mtime is not None:
            self.last_excel_mtime = last_mtime
        return success
             
    def run_debug(self):
        if self.data_handler:
            logging.info("Running dashboard debug mode...")
            self.data_handler.debug_column_mapping()
            self._notify("Column mapping printed to console.", "info")
        else:
            self._notify("No project loaded.", "warning")

    def cleanup_resources(self):
        """Called when app closes."""
        if self._is_cleaned:
            return
        self._is_closing = True
        self._is_cleaned = True

        if self.refresh_timer and self.refresh_timer.isActive():
            self.refresh_timer.stop()
        self._cancel_thumbnail_prefetch()

        self._cleanup_poll_worker(timeout_ms=1500)

        if self.image_loader and hasattr(self.image_loader, "shutdown"):
            try:
                self.image_loader.shutdown(2000)
            except Exception as e:
                logging.debug(f"Image loader shutdown warning: {e}")

        self._cleanup_avatar_upload_worker(timeout_ms=1500)

        if self.file_lock:
            self.file_lock.release()
            self.file_lock = None

    def closeEvent(self, event):
        self._is_closing = True
        self.cleanup_resources()
        event.accept()

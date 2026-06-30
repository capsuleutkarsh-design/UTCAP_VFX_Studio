import logging
import os
import copy
from datetime import datetime
from PySide6.QtWidgets import *
from PySide6.QtCore import *
from PySide6.QtGui import *

from openpyxl import Workbook, load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from ut_vfx.gui.tabs.vfx_dashboard_pro.ui.sync_dialog import SyncDialog
from ut_vfx.gui.tabs.vfx_dashboard_pro.ui.edit_project_dialog import EditProjectDialog

class DashboardActionsMixin:

    def create_blank_template_click(self):
            """Create a blank template Excel using current project structure/mapping."""
            project = self.current_project
            if not project:
                self._notify("Please select a project first.", "warning")
                return

            default_name = f"{project.code}_blank_template.xlsx"
            target_path, _ = QFileDialog.getSaveFileName(
                self,
                "Save Blank Template",
                default_name,
                "Excel Files (*.xlsx)",
            )
            if not target_path:
                return

            try:
                source_path = self.current_excel_path or self.project_manager.get_excel_path(project.code)
                source_wb = None
                source_ws = None
                if source_path and os.path.exists(source_path):
                    source_wb = load_workbook(source_path, data_only=False, keep_vba=False)
                    sheet_name = str(getattr(project, "sheet_name", "") or "").strip()
                    if sheet_name and sheet_name in source_wb.sheetnames:
                        source_ws = source_wb[sheet_name]
                    else:
                        source_ws = source_wb.active

                out_wb = Workbook()
                out_ws = out_wb.active
                out_ws.title = str(getattr(project, "sheet_name", "MASTER") or "MASTER")

                header_row = int(getattr(project, "header_row", 2) or 2)
                data_start_row = int(getattr(project, "data_start_row", 3) or 3)
                mapping = dict(getattr(project, "column_mapping", {}) or {})

                # Preserve top title/meta rows from source when available.
                if source_ws is not None and header_row > 1:
                    for row_idx in range(1, header_row):
                        for col in range(1, source_ws.max_column + 1):
                            out_ws.cell(row=row_idx, column=col, value=source_ws.cell(row=row_idx, column=col).value)
                elif header_row > 1:
                    # Professional title banner when source metadata rows are unavailable.
                    max_banner_col = max(8, max((column_index_from_string(v) for v in mapping.values()), default=8))
                    out_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_banner_col)
                    out_ws.cell(row=1, column=1, value=f"{project.code} - {project.name} | UTVFX Production Template")
                    out_ws.cell(row=1, column=1).font = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
                    out_ws.cell(row=1, column=1).fill = PatternFill(fill_type="solid", fgColor="0F172A")
                    out_ws.cell(row=1, column=1).alignment = Alignment(horizontal="left", vertical="center")
                    out_ws.row_dimensions[1].height = 28

                    if header_row >= 2:
                        out_ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_banner_col)
                        out_ws.cell(row=2, column=1, value="Fill only required columns. Use dropdowns where available.")
                        out_ws.cell(row=2, column=1).font = Font(name="Segoe UI", size=10, italic=True, color="1E293B")
                        out_ws.cell(row=2, column=1).fill = PatternFill(fill_type="solid", fgColor="E2E8F0")
                        out_ws.cell(row=2, column=1).alignment = Alignment(horizontal="left", vertical="center")

                headers = self._build_main_sheet_headers(project, source_ws=source_ws)
                for idx, header in enumerate(headers, start=1):
                    out_ws.cell(row=header_row, column=idx, value=header)

                # Professional styling (inspired by modern dashboard palettes)
                thin = Side(style="thin", color="D1D5DB")
                medium = Side(style="medium", color="334155")
                header_fill_default = PatternFill(fill_type="solid", fgColor="1E3A8A")  # corporate blue
                alt_fill = PatternFill(fill_type="solid", fgColor="F8FAFC")
                section_band_fill = PatternFill(fill_type="solid", fgColor="E0F2FE")
                header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
                body_font = Font(name="Segoe UI", size=10, color="111827")
                center = Alignment(horizontal="center", vertical="center", wrap_text=True)
                left = Alignment(horizontal="left", vertical="center", wrap_text=True)

                # Semantic multi-color header accents.
                palette = {
                    "identity": "1D4ED8",    # Blue
                    "schedule": "0F766E",    # Teal
                    "feedback": "7C3AED",    # Violet
                    "status": "B45309",      # Amber/Brown
                    "artist": "BE185D",      # Pink
                    "delivery": "047857",    # Emerald
                    "department": "6D28D9",  # Deep violet
                    "admin": "374151",       # Slate
                }

                def classify_field(field_name: str) -> str:
                    f = str(field_name or "").lower()
                    if f in {"serial", "thumbnail", "reel", "shot_name", "frames", "sow"}:
                        return "identity"
                    if "feedback" in f or "comment" in f or "note" in f:
                        return "feedback"
                    if "artist" in f:
                        return "artist"
                    if "status" in f:
                        return "status"
                    if any(k in f for k in ("date", "eta", "target", "wip", "done", "submission")):
                        return "schedule"
                    if any(k in f for k in ("version", "priority", "shot_type", "type")):
                        return "admin"
                    if any(k in f for k in ("roto", "dmp", "cg", "prep", "comp", "slapcomp")):
                        return "department"
                    return "delivery"

                field_by_column = {}
                for field_name, col_letter in mapping.items():
                    try:
                        field_by_column[int(column_index_from_string(str(col_letter)))] = str(field_name)
                    except Exception:
                        continue

                # Style main header row
                for idx in range(1, len(headers) + 1):
                    cell = out_ws.cell(row=header_row, column=idx)
                    field_name = field_by_column.get(idx, "")
                    bucket = classify_field(field_name)
                    color = palette.get(bucket, "1E3A8A")
                    cell.fill = PatternFill(fill_type="solid", fgColor=color) if field_name else header_fill_default
                    cell.font = header_font
                    cell.alignment = center
                    cell.border = Border(left=medium, right=medium, top=medium, bottom=medium)

                # Pre-style first 200 blank data rows for production usage.
                preview_rows = 200
                for row_idx in range(data_start_row, data_start_row + preview_rows):
                    is_alt = ((row_idx - data_start_row) % 2) == 1
                    for col_idx in range(1, len(headers) + 1):
                        c = out_ws.cell(row=row_idx, column=col_idx)
                        c.font = body_font
                        c.alignment = left
                        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                        c.protection = Protection(locked=False)
                        if is_alt:
                            c.fill = alt_fill

                # Light section band just beneath header for visual hierarchy.
                band_row = data_start_row
                for col_idx in range(1, len(headers) + 1):
                    if out_ws.cell(row=band_row, column=col_idx).value in (None, ""):
                        out_ws.cell(row=band_row, column=col_idx).fill = section_band_fill

                # Optional formula columns with protection
                serial_col = mapping.get("serial")
                version_col = mapping.get("version")
                latest_version_col = mapping.get("latest_version")
                if serial_col:
                    try:
                        serial_idx = int(column_index_from_string(serial_col))
                        for row_idx in range(data_start_row, data_start_row + preview_rows):
                            cell = out_ws.cell(row=row_idx, column=serial_idx)
                            cell.value = f'=IF($D{row_idx}="","",ROW()-{data_start_row - 1})'
                            cell.alignment = center
                            cell.protection = Protection(locked=True)
                    except Exception:
                        pass

                if latest_version_col and version_col:
                    try:
                        latest_idx = int(column_index_from_string(latest_version_col))
                        version_idx = int(column_index_from_string(version_col))
                        version_letter = get_column_letter(version_idx)
                        for row_idx in range(data_start_row, data_start_row + preview_rows):
                            cell = out_ws.cell(row=row_idx, column=latest_idx)
                            cell.value = f'=IF(${version_letter}{row_idx}="","",${version_letter}{row_idx})'
                            cell.alignment = center
                            cell.protection = Protection(locked=True)
                    except Exception:
                        pass

                if data_start_row > header_row + 1:
                    # Keep expected spacing between header and first data row.
                    out_ws.cell(row=data_start_row - 1, column=1, value=out_ws.cell(row=data_start_row - 1, column=1).value)

                # Carry column widths from source for better usability.
                if source_ws is not None:
                    for idx in range(1, len(headers) + 1):
                        letter = get_column_letter(idx)
                        dim = source_ws.column_dimensions.get(letter)
                        if dim and dim.width:
                            out_ws.column_dimensions[letter].width = dim.width
                        elif out_ws.column_dimensions[letter].width is None or out_ws.column_dimensions[letter].width < 8:
                            out_ws.column_dimensions[letter].width = 16

                # Freeze panes + filter for quick production use.
                out_ws.freeze_panes = f"A{data_start_row}"
                if headers:
                    out_ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{header_row + preview_rows}"
                out_ws.sheet_view.showGridLines = True
                out_ws.row_dimensions[header_row].height = 24
                out_ws.protection.sheet = True
                out_ws.protection.autoFilter = True
                out_ws.protection.sort = True
                out_ws.protection.selectLockedCells = True
                out_ws.protection.selectUnlockedCells = True

                # Data validation dropdowns via hidden lookup sheet
                lookup_ws = out_wb.create_sheet("_LOOKUPS")
                lookups = {
                    "statuses": ["Not Started", "Ready", "WIP", "Review", "Retake", "Final", "Delivered", "Approved"],
                    "priorities": [str(v) for v in (getattr(self.project_manager, "priority_levels", [0, 1, 2, 3]) or [0, 1, 2, 3])],
                    "shot_types": [str(v) for v in (getattr(self.project_manager, "shot_types", ["Prep", "2D Comp", "CG Comp"]) or [])],
                    "artists": sorted({str(u).strip() for u in (self.all_users or []) if str(u).strip()}),
                }
                if not lookups["artists"]:
                    lookups["artists"] = ["Artist_A", "Artist_B"]

                col_cursor = 1
                named_ranges = {}
                for key, values in lookups.items():
                    lookup_ws.cell(row=1, column=col_cursor, value=key.upper())
                    for r, v in enumerate(values, start=2):
                        lookup_ws.cell(row=r, column=col_cursor, value=v)
                    col_letter = get_column_letter(col_cursor)
                    named_ranges[key] = f"'_LOOKUPS'!${col_letter}$2:${col_letter}${len(values)+1}"
                    col_cursor += 1
                lookup_ws.sheet_state = "hidden"

                def add_list_validation(field_names, list_key):
                    if isinstance(field_names, str):
                        field_names = [field_names]
                    for fn in field_names:
                        col_letter = mapping.get(fn)
                        if not col_letter:
                            continue
                        dv = DataValidation(type="list", formula1=f"={named_ranges[list_key]}", allow_blank=True)
                        dv.error = "Select a value from the dropdown list."
                        dv.prompt = "Choose from available values."
                        out_ws.add_data_validation(dv)
                        dv.add(f"{col_letter}{data_start_row}:{col_letter}{data_start_row + preview_rows - 1}")

                add_list_validation(["overall_status", "internal_status", "client_status", "scan_status", "edit_status"], "statuses")
                add_list_validation(["assigned_artist", "artist_all", "roto_artist", "cg_artist", "slapcomp_artist"], "artists")
                add_list_validation("priority", "priorities")
                add_list_validation("shot_type", "shot_types")

                # Print setup
                out_ws.page_setup.orientation = "landscape"
                out_ws.page_setup.fitToWidth = 1
                out_ws.page_setup.fitToHeight = 0
                out_ws.print_title_rows = f"{header_row}:{header_row}"
                out_ws.print_area = (
                    f"A1:{get_column_letter(len(headers))}{header_row + preview_rows}"
                    if headers else None
                )

                # Add auxiliary sheets used by dashboard features.
                extended = getattr(self.project_manager, "extended_sheets", {}) or {}
                for sheet_cfg in extended.values():
                    sheet_name = str(sheet_cfg.get("name", "")).strip()
                    if not sheet_name or sheet_name in out_wb.sheetnames:
                        continue
                    ws = out_wb.create_sheet(sheet_name)
                    cols = dict(sheet_cfg.get("columns", {}) or {})
                    if cols:
                        max_col = 0
                        for letter in cols.values():
                            try:
                                max_col = max(max_col, int(column_index_from_string(str(letter))))
                            except Exception:
                                continue
                        ext_headers = [""] * max_col
                        for key, letter in cols.items():
                            try:
                                pos = int(column_index_from_string(str(letter))) - 1
                            except Exception:
                                continue
                            if 0 <= pos < len(ext_headers):
                                ext_headers[pos] = self._friendly_header_name(key)
                        for idx, header in enumerate(ext_headers, start=1):
                            ws.cell(row=1, column=idx, value=header)
                            hc = ws.cell(row=1, column=idx)
                            hc.fill = PatternFill(fill_type="solid", fgColor="334155")
                            hc.font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
                            hc.alignment = center
                            hc.border = Border(left=medium, right=medium, top=medium, bottom=medium)
                            col_letter = get_column_letter(idx)
                            if ws.column_dimensions[col_letter].width is None or ws.column_dimensions[col_letter].width < 8:
                                ws.column_dimensions[col_letter].width = 18
                        ws.freeze_panes = "A2"

                out_wb.save(target_path)
                if source_wb is not None:
                    source_wb.close()
                self._notify(f"Blank template created: {target_path}", "success", 5000)
            except Exception as e:
                self._notify("Failed to create blank template.", "error", 6000, details=str(e))

    def sync_kitsu_click(self):
            if not self.current_project:
                self._notify("No project loaded to sync.", "warning", 3000)
                return

            self.status_bar.showMessage("Syncing with Kitsu...")
            ok, message, kitsu_shots, conflicts = self.sync_service.sync_with_kitsu(
                self.current_project.name,
                self.all_shots,
            )
            if not ok:
                self._notify(message, "error", 5000)
                return

            resolutions = {}
            if conflicts:
                dialog = SyncDialog(conflicts, self)
                if dialog.exec() != dialog.Accepted:
                    self._notify("Kitsu sync cancelled by user.", "warning", 3500)
                    return
                resolutions = dialog.get_resolutions()

            conflict_ids = {c.get("shot_id") for c in conflicts}
            prefer_cloud_ids = {sid for sid, side in resolutions.items() if side == "cloud"}

            def _merge_shot_from_cloud(local_shot, cloud_shot):
                local_shot.status = cloud_shot.status
                if cloud_shot.description:
                    local_shot.description = cloud_shot.description
                for dept in ["comp_dept", "roto_dept", "prep_dept", "dmp_dept", "cg_dept", "mgfx_dept", "slapcomp_dept"]:
                    cloud_dept = getattr(cloud_shot, dept, None)
                    local_dept = getattr(local_shot, dept, None)
                    if cloud_dept and local_dept and getattr(cloud_dept, "status", ""):
                        local_dept.status = cloud_dept.status

            for k_shot in kitsu_shots:
                match = next((s for s in self.all_shots if s.shot_name == k_shot.shot_name), None)
                if match:
                    if (
                        k_shot.shot_name not in conflict_ids
                        or k_shot.shot_name in prefer_cloud_ids
                    ):
                        _merge_shot_from_cloud(match, k_shot)
                else:
                    self.all_shots.append(k_shot)

            self.save_changes()
            self.update_table()
            self._notify("Kitsu Sync Complete!", "success", 4000)

    def edit_project_click(self):
            if not self.current_project:
                self._notify("Please select a project to edit.", "warning")
                return

            dialog = EditProjectDialog(self.current_project, self)
            if dialog.exec():
                data = dialog.get_data()
                try:
                    success = self.project_manager.update_project(
                        data['code'],
                        data['name'],
                        data['excel_path'],
                        data['folder_base'],
                        sheet_name=data.get('sheet_name'),
                        header_row=data.get('header_row'),
                        data_start_row=data.get('data_start_row')
                    )
                    if success:
                        self._notify("Project configuration updated successfully.", "success")
                        # Refresh
                        self.load_projects()
                        # Re-select same project to reload config
                        idx = self.project_combo.findData(data['code'])
                        if idx >= 0:
                            self.project_combo.setCurrentIndex(idx)
                            self.switch_project(data['code'])
                    else:
                        self._notify("Failed to update project.", "warning")
                except Exception as e:
                    self._notify("Project update failed.", "error", details=str(e))

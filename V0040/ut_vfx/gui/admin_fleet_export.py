"""
Fleet report export helpers for Admin Panel.
"""

from datetime import datetime


def export_fleet_xlsx(output_path, records, summary, skipped):
    """Create a color-coded, bordered Excel workbook for the fleet report."""
    from openpyxl import Workbook
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # Palette
    col_header_bg = "1E2A3A"
    col_header_fg = "FFFFFF"
    col_online_bg = "D4EDDA"
    col_online_fg = "155724"
    col_idle_bg = "FFF3CD"
    col_idle_fg = "856404"
    col_offline_bg = "F8D7DA"
    col_offline_fg = "721C24"
    col_unknown_bg = "E2E3E5"
    col_unknown_fg = "383D41"
    col_crit_bg = "FF0000"
    col_warn_bg = "FFA500"
    col_ok_bg = "28A745"
    col_summary_bg = "2C3E50"
    col_tile_online = "27AE60"
    col_tile_idle = "F39C12"
    col_tile_off = "E74C3C"

    thin = Side(style="thin", color="CCCCCC")
    Side(style="medium", color="999999")
    thick = Side(style="thick", color=col_header_bg)
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    def make_fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def make_font(hex_color, bold=False, size=10):
        return Font(color=hex_color, bold=bold, size=size, name="Segoe UI")

    # Sheet 1: Summary
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.sheet_view.showGridLines = False

    ws_sum.merge_cells("A1:F1")
    title_cell = ws_sum["A1"]
    title_cell.value = f"UT_VFX Fleet Report - {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    title_cell.font = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
    title_cell.fill = make_fill(col_summary_bg)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[1].height = 36

    tiles = [
        ("Total Workstations", len(records), "3498DB"),
        ("Online", summary["online"], col_tile_online),
        ("Idle", summary["idle"], col_tile_idle),
        ("Offline", summary["offline"], col_tile_off),
        ("Skipped (errors)", skipped, "95A5A6"),
    ]
    headers_row, values_row = 3, 4
    for col_idx, (label, val, color) in enumerate(tiles, start=1):
        h_cell = ws_sum.cell(row=headers_row, column=col_idx, value=label)
        h_cell.fill = make_fill(color)
        h_cell.font = make_font("FFFFFF", bold=True, size=9)
        h_cell.alignment = Alignment(horizontal="center", vertical="center")
        h_cell.border = border_thin
        ws_sum.row_dimensions[headers_row].height = 20

        v_cell = ws_sum.cell(row=values_row, column=col_idx, value=val)
        v_cell.fill = make_fill("F8F9FA")
        v_cell.font = Font(name="Segoe UI", size=22, bold=True, color=color)
        v_cell.alignment = Alignment(horizontal="center", vertical="center")
        v_cell.border = border_thin
        ws_sum.row_dimensions[values_row].height = 44
        ws_sum.column_dimensions[get_column_letter(col_idx)].width = 22

    # Sheet 2: Fleet data
    ws = wb.create_sheet("Fleet Data")
    ws.sheet_view.showGridLines = False

    if not records:
        ws["A1"] = "No data found."
        wb.save(output_path)
        return

    all_keys = list(dict.fromkeys(k for r in records for k in r))

    header_map = {
        "pc_name": "PC Name", "status": "Status", "last_seen": "Last Seen",
        "last_seen_age": "Age", "age_seconds": "Age (s)",
        "ut_user": "UT User", "os_user": "OS User",
        "ip_address": "IP Address", "mac_address": "MAC Address",
        "computer_name": "Computer Name",
        "manufacturer": "Manufacturer", "model": "Model",
        "motherboard": "Motherboard", "serial_no": "Serial No",
        "cpu": "CPU", "gpu": "GPU", "ram_gb": "RAM",
        "os": "OS", "windows_version": "Win Version",
        "client_version": "Client Ver",
        "disk_c_total_gb": "C: Total GB", "disk_c_free_gb": "C: Free GB",
        "disk_c_usage": "C: Usage", "disk_c_alert": "C: Alert",
    }

    header_fill = make_fill(col_header_bg)
    header_font = make_font(col_header_fg, bold=True, size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_border = Border(left=thin, right=thin, top=thick, bottom=thick)

    for col_idx, key in enumerate(all_keys, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header_map.get(key, key.replace("_", " ").title()))
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = header_border
    ws.row_dimensions[1].height = 30

    status_styles = {
        "Online": (col_online_bg, col_online_fg),
        "Idle": (col_idle_bg, col_idle_fg),
        "Offline": (col_offline_bg, col_offline_fg),
        "Unknown": (col_unknown_bg, col_unknown_fg),
    }
    alert_fills = {
        "CRITICAL": make_fill(col_crit_bg),
        "WARNING": make_fill(col_warn_bg),
        "OK": make_fill(col_ok_bg),
    }
    alert_fonts = {
        "CRITICAL": make_font("FFFFFF", bold=True),
        "WARNING": make_font("000000", bold=True),
        "OK": make_font("FFFFFF", bold=True),
    }

    divider_border = Border(left=thin, right=thin, top=thick, bottom=thick)
    status_order = {"Online": 0, "Idle": 1, "Offline": 2, "Unknown": 3}
    divider_labels = {0: "  ONLINE", 1: "  IDLE", 2: "  OFFLINE", 3: "  UNKNOWN"}
    divider_colors = {0: "27AE60", 1: "F39C12", 2: "E74C3C", 3: "95A5A6"}

    row_idx = 2
    prev_group = -1
    for record in records:
        status = record.get("status", "Unknown")
        group = status_order.get(status, 3)

        if group != prev_group:
            d_fill = make_fill(divider_colors.get(group, "95A5A6"))
            first_cell = ws.cell(row=row_idx, column=1, value=divider_labels.get(group, ""))
            first_cell.fill = d_fill
            first_cell.font = Font(name="Segoe UI", size=9, bold=True, color="FFFFFF", italic=True)
            first_cell.border = divider_border
            for col_idx in range(2, len(all_keys) + 1):
                dc = ws.cell(row=row_idx, column=col_idx, value="")
                dc.fill = d_fill
                dc.border = divider_border
            ws.row_dimensions[row_idx].height = 14
            row_idx += 1
            prev_group = group

        row_fill = make_fill(status_styles.get(status, (col_unknown_bg, col_unknown_fg))[0])
        row_font = make_font(status_styles.get(status, (col_unknown_bg, col_unknown_fg))[1])

        for col_idx, key in enumerate(all_keys, start=1):
            val = record.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = row_fill
            cell.font = row_font
            cell.border = border_thin
            cell.alignment = Alignment(vertical="center")

            if key == "status":
                cell.font = make_font(status_styles.get(status, (col_unknown_bg, col_unknown_fg))[1], bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")

            if key.endswith("_alert") and val in alert_fills:
                cell.fill = alert_fills[val]
                cell.font = alert_fonts[val]
                cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[row_idx].height = 18
        row_idx += 1

    for col_idx, key in enumerate(all_keys, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(header_map.get(key, key)),
            max((len(str(r.get(key, ""))) for r in records), default=0),
        )
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 40)

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(all_keys))}1"

    wb.save(output_path)
